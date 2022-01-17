#======================================================================================================================================================
#======================================================================================================================================================

# Imports

import os, configparser, json, OpenSSL


from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from tabulate import tabulate

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


#======================================================================================================================================================
#======================================================================================================================================================

# Variables

msg = MIMEMultipart()
formatter = '\n' + "____________________________________________________________________________________" + "\n"
Occurances_Summary=[]
Files_List=[]
Search_Dirs=[]
#Search_Dirs=["\\var\\log\\firemon\\nd", "\\var\\log\\fmos"]
#Search_Dirs=["\\var\\log\\firemon\\sm\\secmgr.log"]
Health_File_Name='health.log'
ssl_access_File_Name='ssl_access_log'
ndexec_File_Name='ndexec'
directory_search_output=[]
directory_search_output_analysis=[]
es_File_Name='es.log'
secmgr_File_Name='secmgr'
wf_File_Name='wf.log'
messages_file_name='messages'
ssl_access_Search_Criteria=[]
ndexec_Search_Criteria=[]
Health_Search_Criteria=[]
es_Search_Criteria=[]
wf_Search_Criteria=[]
secmgr_Search_Criteria=[]
messages_Search_Criteria=[]
Top_Largest_files=[] 
global_search=[]
Percentage_of_Licences_Allocated=[]
Days_Until_Certificates_Expire=[]
Device_Uptime=30

#Device Inventory Report
Device_Count=[]
device_type=[]
Device_Count2=[]
device_type2=[]
Check_Installed_DP=0
Total_Devices_Summary=""
Check_Installed_DPs=""
DP_list=[]

allFiles=[]
biggest2=[]
biggest3=[]
alldirs=[]


messagesDropped_list=[]
messageParsingErrors_list=[]
Drops_Found=0
Drops_Summary=""


final_unread=[]
unread=[]
all_files_full_path=[]
global_search_count=0

summary_cpu=[]
summary_cpu_count=0

exception_list=[]

lic_max_util=0
total_finds=[]


certs='\\etc\\pki\\tls\\certs\\localhost.crt', '\\etc\\pki\\ca-trust\\source\\anchors\\fmos-root.crt', '\\etc\\pki\\tls\\certs\\fmos-admin.cer'
display_name=''
warn_expire=0
Summary_Dir_Check='WARNING -- .xy files skipped'

cpu_count=0
temp_files_count=0

#======================================================================================================================================================
#======================================================================================================================================================

# Reading the Config File and storing contents

try:


	config = configparser.ConfigParser()
	config.read("mkdiag_script_config_v1.ini")


	for key in config['health']:
		value=config['health'][key]
		Health_Search_Criteria.append(value)


	for key in config['ndexec']:
		value=config['ndexec'][key]
		ndexec_Search_Criteria.append(value)


	for key in config['ssl_access']:
		value=config['ssl_access'][key]
		ssl_access_Search_Criteria.append(value)
		

	for key in config['No_Of_Files']:
		value=config['No_Of_Files'][key]
		Top_Largest_files.append(value)


	for key in config['Global_Search']:
		value=config['Global_Search'][key]
		global_search.append(value)


	for key in config['Percentage_of_Licences_Allocated']:
		value=config['Percentage_of_Licences_Allocated'][key]
		Percentage_of_Licences_Allocated.append(value)


	for key in config['Days_Until_Certificates_Expire']:
		value=config['Days_Until_Certificates_Expire'][key]
		Days_Until_Certificates_Expire.append(value)


	for key in config['Specific_Search_Dirs']:
		value=config['Specific_Search_Dirs'][key]
		Search_Dirs.append(value)


	for key in config['Device_Uptime_Config']:
		value=config['Device_Uptime_Config'][key]
		Device_Uptime=value


	for key in config['es']:
		value=config['es'][key]
		es_Search_Criteria.append(value)


	for key in config['messages']:
		value=config['messages'][key]
		messages_Search_Criteria.append(value)


	for key in config['secmgr']:
		value=config['secmgr'][key]
		secmgr_Search_Criteria.append(value)


	for key in config['wf']:
		value=config['wf'][key]
		wf_Search_Criteria.append(value)



	# for key in config['Days_Until_Certificates_Expire']:
	# 	value=config['Days_Until_Certificates_Expire'][key]
	# 	Days_Until_Certificates_Expire.append(value)

		


except Exception as e:
	print(e)
	print("Error reading config file")
	exception_list.append(e)



#======================================================================================================================================================
#======================================================================================================================================================

# Functions

def directory_search(MyLoc: object, Dir: object, Log: object, words: object) -> object:

	#Get a list of all files in a directory
	dir_search = os.listdir( MyLoc + Dir)
	Files_List=[]

	#Filter the list provided above to only files that contain string of interest
	for item_nw in dir_search:
		#print(item_nw)
		if Log in item_nw: #Pass in this string in future
			Files_List.append(item_nw)
		else:
			pass

	#Loop through each log file of interest, reading each 
	for Log_Check in Files_List:
		#print(Log_Check)
		if '.xz' in Log_Check:
			Summary_Dir_Check='WARNING -- .xy files skipped'
			pass
		else:
			Summary_Dir_Check='PASS -- .xy files not skipped'
			f = open(str(windows_pwd) + Dir +"\\" + str(Log_Check))
			lines = f.read()
			occurances=0
			
			#Grab a word to search on, count total occurances in file and add to Occurances_Summary
			#for word in words:
			occurances=lines.count(words)
			Occurances_Summary.append(str(Log_Check)+ " --- "+str(words)+' --- '+str(occurances))
			

			#Divide the file into lines and check if it contains the word of interest	
			for line in lines.split('\n'):
				if words in line:

					#Output successful matches to various places 
					print(Log_Check + " ---- "+ line + "\n")
					directory_search_output.append(line)
					#Health_LoggingFile.write(Log_Check + " ---- "+ line + "\n")
				else:
					pass



#======================================================================================================================================================
#======================================================================================================================================================

# Other

pwd=os.getcwd()
print("Present Working Directory " + pwd)
windows_pwd=pwd.replace('\\','\\\\')

try:
	os.mkdir(pwd + "\\script_output")
except:
	pass


wb=load_workbook('formatted.xlsx')
ws = wb.active
ws.title='nicko'


Summary_LoggingFile = open(pwd + "\\script_output\\Summary_Logging_" + os.path.basename(__file__) + ".txt" , "w")


#======================================================================================================================================================
#======================================================================================================================================================

# mkdiagpkg.out

print(formatter)

print("Opening mkdiagpkg.out")
print()
f = open("mkdiagpkg.out")
mkdiagpkg_out = f.read().splitlines()


FMOSv=mkdiagpkg_out[1].split("FMOS release ")
mkdiag_created=mkdiagpkg_out[2]
hostname=mkdiagpkg_out[3].split(" ")[1]
Device_Uptime_host=(mkdiagpkg_out[4].split("up")[1].split(",")[0])
Device_Uptime_host = Device_Uptime_host.strip().split(' ')[0]


if 'min' in Device_Uptime_host:
	Device_Uptime_host_days=1
else:
	Device_Uptime_host_days=Device_Uptime_host




FMOSv=FMOSv[1]
short_FMOSv = FMOSv[:-2]



print("MKDIAG Generated: " + str(mkdiag_created))
print("FMOS Hostname: " + str(hostname))
print("FMOS Version: " + str(FMOSv))
print("Uptime: " + str(Device_Uptime_host_days) + " day(s)")
print()
if int(Device_Uptime_host) > int(Device_Uptime):
	print("WARNING -- Device Uptime is greater than " + str(Device_Uptime)+ " days.")
	Check_Device_Uptime="WARNING -- Device Uptime is greater than " + str(Device_Uptime)+ " days."
else:
	print("PASS -- Device Uptime is less than " + str(Device_Uptime)+ " days.")
	Check_Device_Uptime="PASS -- Device Uptime is less than " + str(Device_Uptime)+ " days."

print()
print("Closing mkdiagpkg.out")
f.close()
print(formatter)





#======================================================================================================================================================
#======================================================================================================================================================

# Device Inventory Report

print("Opening deviceInventoryReport")
print()
try:
	if os.path.exists(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\deviceInventoryReport.txt"):

		f = open(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\deviceInventoryReport.txt", encoding="utf8")
		deviceInventoryReport = f.readlines()

		for line in deviceInventoryReport:
			if "Device" in line:
				Device_Count.append(line)
			if 'DevicePack Artifact ID' in line:

				device_type.append(line)

		for item in Device_Count:
			if "/" in item:
				Device_Count2.append(item)
			if "Version" in item:
				Device_Count2.append(item)
				DP_Ver_Check=item.split(".")
				Storing_DP=DP_Ver_Check[0][-1]+'.'+ DP_Ver_Check[1][0]+'.'+ DP_Ver_Check[2][:2]
				DP_Ver_Check=DP_Ver_Check[0][-1]+'.'+ DP_Ver_Check[1]
				if DP_Ver_Check!= short_FMOSv:
					Check_Installed_DP+=1
				Storing_DP = Storing_DP.strip()
				if Storing_DP not in  DP_list:
					DP_list.append(Storing_DP)

		
		for item in device_type:

			item=item.split("│")[2].strip()

			device_type2.append(item)
		#print(device_type2)
		print()

		


		"""

		Make nice list showing device count

		"""
		device_totals=[]
		device_type3 = list(set(device_type2))
		# for line in device_type2:
		# 	print(line)


		for item in device_type3:		
			x=device_type2.count(item)
			emptylist=[] 		
			emptylist.append(item)
			emptylist.append(x)
			device_totals.append(emptylist)


		device_totals.sort(key = lambda x: x[1])
		print(tabulate(device_totals, headers=['Device', 'Count']))


		"""

		End of Make nice list showing device count

		"""


		print()
		Total_Devices=len(Device_Count2)/2
		print("There are "+str(int(Total_Devices))+" devices, using the following Device Packs")
		Total_Devices_Summary="There are "+str(int(Total_Devices))+" devices"
		

		print(', '.join(DP_list))
		print()

		if Check_Installed_DP != 0:
			print("WARNING -- Check Installed Device Packs")
			Check_Installed_DPs="WARNING -- Check Installed Device Packs"
		else:
			print("PASS -- Appropiate Device Packs Installed")
			Check_Installed_DPs="PASS -- Appropiate Device Packs Installed"
	else:
		Check_Installed_DPs="N/A -- Device Packs Check not performed. No DC element to MKDIAG"
		print(Check_Installed_DPs)






		
	print()
	print("Closing deviceInventoryReport")
	f.close()
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking Device Inventory")
	print()
	print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Directory Searches


print("Begining Directory Searches")
print()


if len(Health_Search_Criteria)==0 and len(ssl_access_Search_Criteria) ==0 and len(ndexec_Search_Criteria)==0:
	print('---- No Directory Searches Requested ----')


#print(Search_Dirs)


for Search_Dir in Search_Dirs:
	#print(Search_Dir)
	if 'fmos' in Search_Dir:
		DirF=Search_Dir
		LogF=Health_File_Name
		WordsF=Health_Search_Criteria
	elif 'httpd' in Search_Dir:
		DirF=Search_Dir
		LogF=ssl_access_File_Name
		WordsF=ssl_access_Search_Criteria
	elif 'nd' in Search_Dir:
		DirF=Search_Dir
		LogF=ndexec_File_Name
		WordsF=ndexec_Search_Criteria
	elif 'wf' in Search_Dir:
		DirF=Search_Dir
		LogF=wf_File_Name
		WordsF=wf_Search_Criteria
	elif 'sm' in Search_Dir:
		DirF=Search_Dir
		LogF=secmgr_File_Name
		WordsF=secmgr_Search_Criteria
	elif 'elasticsearch' in Search_Dir:
		DirF=Search_Dir
		LogF=es_File_Name
		WordsF=es_Search_Criteria
	elif 'messages' in Search_Dir:
		DirF=Search_Dir
		LogF=messages_file_name
		WordsF=messages_Search_Criteria
	# elif 'nd' in Search_Dir:
	# 	DirF=Search_Dir
	# 	LogF=ndexec_File_Name
	# 	WordsF=ndexec_Search_Criteria





	print(WordsF)
	for Words101 in WordsF:
		print(Words101)
		directory_search(str(windows_pwd), DirF, LogF, Words101)


print()
print("Completed Directory Searches")
print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Top 3 biggest files

print("Calculating the 3 largest files present")
print()
try:

	for root, dirs, files in os.walk(".", topdown=True):
		for name in files:
			full_path=os.path.join(root, name)
			allFiles.append(full_path)


	max_file = max(allFiles, key =  lambda x: os.stat(x).st_size)
	print('Largest File (GB\'s): ', os.stat(max_file).st_size/1000000000)
	print(max_file)


	for item in allFiles:
		if str(max_file) in item:
			pass
		else:
			biggest2.append(item)



	max_file2 = max(biggest2, key =  lambda x: os.stat(x).st_size)
	print()
	print('Second Largest File (GB\'s): ', os.stat(max_file2).st_size/1000000000)
	print(max_file2)


	for item in biggest2:
		if str(max_file2) in item:
			pass
		else:
			biggest3.append(item)

	max_file3 = max(biggest3, key =  lambda x: os.stat(x).st_size)
	print()
	print('Third Largest File (GB\'s): ', os.stat(max_file3).st_size/1000000000)
	print(max_file3)


except Exception as e:
	print(e)
	print("Error Calculating largest 3 files")


print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Metrics Log

try:

	print("Opening metrics.log")
	print()
	if os.path.exists(str(windows_pwd) +"\\var\\log\\firemon\\dc\\metrics.log"):


		f = open(str(windows_pwd) +"\\var\\log\\firemon\\dc\\metrics.log")
		metrics = f.readlines()

		for line in metrics:
			if "messagesDropped" in line:
				messagesDropped_list.append(line)
			elif "messageParsingErrors" in line:
				messageParsingErrors_list.append(line)
			else:
				pass
		
		bad_syslog_summary = "PASS -- No Evidence Of Bad Syslog Found"
		for item in messageParsingErrors_list:
			if '0' not in item:
				bad_syslog_summary = "WARNING -- Evidence Of Bad Syslog Found - Check \\var\\log\\firemon\\dc\\metrics.log"

		print(bad_syslog_summary)

		for item in messagesDropped_list:
			item=item.split(':')
			clean_item=item[1]
			clean_item=clean_item.strip()
			number_check=int(clean_item)
			if number_check > 0:
				print(item)
				Drops_Found+=1
			else:
				pass



		if Drops_Found > 0:
			Drops_Summary="WARNING -- Evidence Of Dropped Packets Found - Check \\var\\log\\firemon\\dc\\metrics.log"
		else:
			Drops_Summary="PASS -- No Evidence Of Dropped Packets Found"

		print(Drops_Summary)

	else:
		Drops_Summary="N/A -- Dropped Packets Check not performed. No DC element to MKDIAG"
		print(Drops_Summary)
		bad_syslog_summary="N/A -- Bad Syslog Check not performed. No DC element to MKDIAG"
		print(bad_syslog_summary)


	print()
	print("Closing metrics.log")

except Exception as e:
	print(e)
	print("Error Checking metrics.log")

print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Global Search

try:
	print("Start of performing Global Search")
	print()



	if len(global_search) > 0:
		print(', '.join(global_search))



		for gs in global_search:
			print('Looking for '+str(gs))
			print()
			for dirpath, subdirs, files in os.walk(".", topdown=True):
			    for x in files:
			    	all_files_full_path.append(os.path.join(dirpath, x))



			for name in all_files_full_path:
				if 'mkdiag_script_config_v1.ini'in name:
					pass
				else:
					try:

						f = open(name)
						lines = f.readlines()
						for line in lines:
							if gs in line:
								global_search_count=global_search_count+1
								print(name)
								total_finds.append(name)
								print(line)

					except:
						unread.append(name)



			i=1
			while i < 10:
				i=i+1
				for unreadable in unread:
					try:
						f = open(unreadable, encoding="utf8")
						fc = f.readlines()
						for line in lines:
							if gs in line:
								global_search_count=global_search_count+1
								print(line)
						unread.remove(unreadable)
					except:
						pass
		print()
		print()
		total_finds = list(dict.fromkeys(total_finds))
		print(', '.join(total_finds))
		print()
		print()
	
	else:
		print('---- No Global Search Requested ----')
		print()

	print("End of performing Global Search")

except Exception as e:
	print(e)
	print("Error performing Global Search")


print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Top

try:

	print("Opening Top")
	print()
	f = open(str(windows_pwd) +"\\top.txt")
	top = f.read()
	line = top.split('\n')

	cpu_loadavg_1=line[0].split('average: ')[1].split(',')[0]
	cpu_loadavg_5=line[0].split('average: ')[1].split(',')[1].strip()
	cpu_loadavg_15=line[0].split('average: ')[1].split(',')[2].strip()

	MiB_Mem_clean=line[3]
	MiB_Swap_clean=line[4]

	# print(MiB_Mem_clean)
	# print(MiB_Swap_clean)

	MiB_Mem_free=MiB_Mem_clean.split(" free")[0]
	MiB_Mem_free=MiB_Mem_free.split(" ")[-1]
	MiB_Mem_free=int(float(MiB_Mem_free))


	MiB_Mem_total=MiB_Mem_clean.split(" total")[0]	
	MiB_Mem_total=MiB_Mem_total.split(" ")[-1]
	MiB_Mem_total=int(float(MiB_Mem_total))



	mem_total_used = int(float(MiB_Mem_total - MiB_Mem_free))
	mem_total_used_gb=round(mem_total_used/1000)
	mem_pc=float(mem_total_used / MiB_Mem_total ) *100
	print('The total physical memory in use is '+str(mem_total_used_gb)+'GB, that\'s approx '+str(int(mem_pc)) +'%.')
	print()

	top_memory_summary="PASS -- Less Than 90% Of Total Memory In Use"
	
	if mem_pc > 90:
	    top_memory_summary="WARNING -- Greater Than 90% Of Total Memory In Use"
	print(top_memory_summary)


	print()

	#Swap Mem

	MiB_Swap_used=MiB_Swap_clean.split(" used")[0]
	MiB_Swap_used=MiB_Swap_used.split(" ")[-1]
	MiB_Swap_used=(float(MiB_Swap_used)/1000)
	MiB_Swap_used=round(MiB_Swap_used)


	top_swap_summary= "PASS -- Swap Memory Not In Use"
	if MiB_Swap_used > 0:
	    top_swap_summary="WARNING -- Swap Memory In Use, Approx " +str(MiB_Swap_used)+'GB'
	print(top_swap_summary)

	print()

	#CPU
	cpu_des=line[6]

	newlist=[]
	newlist.append(line[7:])

	for item in newlist[0]:
	    all_ent=[]
	    item2=item.split(" ")
	    for items in item2:
	        if items != "":
	            all_ent.append(items)
	    try:
	        cpu_usage=int(float(all_ent[8]))
	    except:
	        pass



	    if cpu_usage > 85:
	        summary_cpu.append(item)
	        summary_cpu_count+=1
	        #print(item)
	    else:
	        pass

	if summary_cpu_count > 0:
	    summary_cpu_text="WARNING -- The Processes Are Using 85% CPU Or More"
	    print(summary_cpu_text)
	    print(cpu_des)
	    for process in summary_cpu:
	        print(process)
	else:
		summary_cpu_text="PASS -- Top Does Not Show Excessive CPU Usage"
		print(summary_cpu_text)
	print()
	print("Closing Top")


except Exception as e:
	print(e)
	print()
	print("Error reading Top")

print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# Crashdump

print("Checking for Crash Dumps")
print()
try:


	os.chdir(str(windows_pwd) +"\\var\\lib")
	if os.path.isdir('crashdump'):
		os.chdir(str(windows_pwd) +"\\var\\lib\\crashdump")

		here=os.listdir()

		if len(here) > 0:

			Core_Summary='WARNING -- Core Files Found'
			print(Core_Summary)
			for core in here:
				print(core)
		else:
			Core_Summary='PASS -- No Core Files Found'
			print(Core_Summary)

	else:
		Core_Summary='PASS -- No Core Files Found'
		print(Core_Summary)

	os.chdir(str(windows_pwd))


		
	print()
	print("Closing Check for Crash Dumps")
	f.close()
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking for Crash Dumps")



#======================================================================================================================================================
#======================================================================================================================================================

# Off Site Backup

print("Checking for Off Box Backup")
print()
try:


	os.chdir(str(windows_pwd) +"\\etc\\firemon\\postbackup.d")
	here=os.listdir()


	if len(here) < 3:

		Backup_Summary='WARNING -- Check if \'Off Box\' Backup Configured'
		print(Backup_Summary)
		print()
		for her in here:
			print(her)

	else:
		Backup_Summary='PASS -- \'Off Box\' Backup Configured'
		print(Backup_Summary)


	os.chdir(str(windows_pwd))

		
	print()
	print("Closing Check for Off Box Backup")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking for Off Box Backup")



#======================================================================================================================================================
#======================================================================================================================================================

# sm_diagpkg.json 

print("Beginning to parse sm-diagpkg.json")
print()



try:
	License_Allocation=int(Percentage_of_Licences_Allocated[0])


	r=open('sm-diagpkg.json')
	f = json.load(r)
	counter=0
	text=""
	Norm_Ret_counter=0

	comp_name=str(f["companyName"])
	device_count=str(f["deviceCount"])

	print("Summary for: "+str(comp_name))
	print("Total Devices: "+str(device_count))
	print()
	print()
	print('Using ' +str(License_Allocation)+ '% allocation and greater to fire warning')
	for n in f['domains'][0]['licenseAllocations']:
		categoryName=n['categoryName']
		licenseTotal=n['licenseTotal']
		licenseUsed=n['licenseUsed']


		if licenseTotal!=0:


			print('License Type: '+str(categoryName))
			print('Total Available '+str(licenseTotal))
			utilization=licenseUsed/licenseTotal*100
			if utilization > License_Allocation:
				lic_max_util+=1
			print('Currently Used '+str(int(utilization))+ "%")
			print()

	if lic_max_util > 0:
		License_Summary= 'WARNING -- A License Allocation Over '+str(License_Allocation)+'% Threshold'
		print(License_Summary)
	else:
		License_Summary= 'PASS -- All License Allocations Under '+str(License_Allocation)+'% Threshold'
		print(License_Summary)
	print()


	for n in f['domains'][0]['devices']:
		revisionlist=[]
		counter+=1
		name=n['name']
		retrievalError=n['retrievalError']
		normalizationError=n['normalizationError']
		devicePack=n['devicePack']['version']
		for revision in n['revisionList']:
			id=revision['id']
			revisionlist.append(id)

		if normalizationError == 0 and retrievalError == 0:
			#text="There are no known retrieval or normalization errors for this device"
			pass
		else:

			if retrievalError != 0:
				Norm_Ret_counter+=1
				text="WARNING - There are RETRIEVAL errors for this device"
			if normalizationError != 0:
				Norm_Ret_counter+=1
				text="WARNING - There are NORMALIZATION errors for this device"
			if normalizationError > 0 and retrievalError > 0:
				Norm_Ret_counter+=1
				text="There are BOTH RETRIEVAL AND NORMALIZATION errors for this device"




			print('Device Name: '+str(name))
			print('Device Pack: '+str(devicePack))
			revisionlist.sort()
			print('Known revisions are: '+str(revisionlist))
			print(text)
			print()
	

	if Norm_Ret_counter > 0:
		Norm_Ret_Summary='WARNING -- NORMALIZATION / RETRIEVAL Errors Found (sm-diagpkg.json)'
		print(Norm_Ret_Summary)
	else:
		Norm_Ret_Summary='PASS -- No NORMALIZATION / RETRIEVAL Errors Found (sm-diagpkg.json)'
		print(Norm_Ret_Summary)



	print()
	print("Finishing parsing sm-diagpkg.json")
	
	print(formatter)

except Exception as e:
	print(e)
	print("Error parsing sm-diagpkg.json")
	print()
	print(formatter)


#======================================================================================================================================================
#======================================================================================================================================================

# SSL Certificates 

print("Checking Certificate Expiry Dates")
print()

days_till_expiry=int(Days_Until_Certificates_Expire[0])



try:

	for cert in certs:
		cert_name=cert.split('\\')[-1]
		if cert_name =='localhost.crt':
			display_name = 'Server (Apache / HTTPS)'
		elif cert_name =='fmos-root.crt':
			display_name = 'FMOS Ecosystem Root CA'
		elif cert_name =='fmos-admin.cer':
			display_name = 'FMOS Control Panel'
		try:
			with open(windows_pwd +cert, "r") as my_cert_file:
					my_cert_text = my_cert_file.read()
					cert = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, my_cert_text)



					exp_date=datetime.strptime(cert.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')
					#exp_date2=datetime.strptime(exp_date, '%Y-%m-%d')
					today=datetime.today().strftime('%Y-%m-%d')


					print(cert_name)
					print(display_name)
					print('Expires: '+str(exp_date))
					delta = exp_date - datetime.today()
					print('Days till expiry: '+str(delta.days))
					if days_till_expiry >= delta.days:
						warn_expire+=1
					print()
		except:
			pass


	if warn_expire > 0:
		Sumarry_Cert=('WARNING -- A Certificate Expires Soon')
		print(Sumarry_Cert)
		print()
	else:
		Sumarry_Cert=('PASS -- Certificate Expiry Date Greater Than '+ str(days_till_expiry) +' Days')
		print(Sumarry_Cert)
		print()
		print()





except Exception as e:
	print(e)
	print("Error Checking Certificate Expiry Dates")
	print()
	print(formatter)

print("Closing Check for Certificate Expiry Dates")

print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# Device Role

print("Checking for device Role")
print()
try:

	f = open(str(windows_pwd) +"\\etc\\firemon\\fm_roles")
	fm_role = f.readlines()

	if len(fm_role) == 1:
		device="DC only"



	if '0' in fm_role[0]  and '1' in fm_role[1] and '1' in fm_role[2] and '1' in fm_role[3] and '1' in fm_role[4]:
		device="AS and DB"
	elif '0' in fm_role[0]  and '1' in fm_role[1] and '0' in fm_role[2] and '0' in fm_role[3] and '0' in fm_role[4]:
		device="AS only"
	elif '0' in fm_role[0]  and '0' in fm_role[1] and '1' in fm_role[2] and '0' in fm_role[3] and '0' in fm_role[4]:
		device="DB only"
	elif '1' in fm_role[0]  and '1' in fm_role[1] and '1' in fm_role[2] and '1' in fm_role[3] and '1' in fm_role[4]:
		device="All in One"



	print(device)
	

	print()
	print("Closing Check for device Role")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking for for device Role")
	print()
	print(formatter)



#======================================================================================================================================================
#======================================================================================================================================================

# CPU INFO

print("Checking CPU's")
print()
try:

	f = open(str(windows_pwd) +"\\cpuinfo.txt")
	cpu_info = f.readlines()



	for line in cpu_info:
		if 'processor	: ' in line:
			cpu_count += 1
	print('No of CPUs found '+str(cpu_count))
	print()
	print()

	cpuload=[['1',cpu_loadavg_1],['5',cpu_loadavg_5],['15',cpu_loadavg_15]]
	print(tabulate(cpuload, headers=['Time Mins', 'CPU Load Avg']))

	
	cpu_info_xls=[cpu_count,cpu_loadavg_1,cpu_loadavg_5,cpu_loadavg_15]
	# print(cpu_info_xls)

	print()
	print("Closing Check of CPU's")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking CPU's")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================



#======================================================================================================================================================
#======================================================================================================================================================

# dc.conf

print("Checking dc.conf file")
print()
try:


	f = open(str(windows_pwd) +"\\etc\\firemon\\dc.conf")
	dc_conf = f.readlines()
	for line in dc_conf:
		if "--DataCollector.SyslogServer.ThreadNumberForProcessingMessages" in line:
			print(line)
			line=line.split(" ")[1]
			cpu_minus_one = cpu_count-1
			if line != cpu_minus_one:
				Sumarry_CPU=('WARNING -- Review Thread Number For Processing Messages')
				print(Sumarry_CPU)
			else:
				Sumarry_CPU=('PASS -- Thread Number For Processing Messages is set correctly')
				print(Sumarry_CPU)
	

	print()
	print("Closing Check of dc.conf file")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking dc.conf file")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# Java heap size

print("Checking Java heap size")
print()
try:
	if os.path.exists(str(windows_pwd) +"\\etc\\firemon\\sm.jvm.options"):

		r=open(str(windows_pwd) +"\\etc\\firemon\\sm.jvm.options")
		s=open(str(windows_pwd) +"\\etc\\firemon\\wf.jvm.options")
		t=open(str(windows_pwd) +"\\etc\\firemon\\nd.jvm.options")

		
		sm = r.readlines()
		nd = s.readlines()
		wf = t.readlines()

		print('SM '+str(sm[3].split('mx')[1]))
		print('WF '+str(wf[3].split('mx')[1]))
		print('ND '+str(nd[3].split('mx')[1]))
		print()
		
		print(formatter)

	else:
		print('This device is a ' + str(device) + ' therefore no Java heap size files present')
		print()
		print("Closing Check for Java heap size")
		print(formatter)

except Exception as e:
	print(e)
	print("Error Checking Java heap size")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# Detailed Health check

print("Checking Health")
print()

DirF="\\var\\log\\fmos"
LogF=Health_File_Name
WordsF2=Health_Search_Criteria

for WordsF in WordsF2:
	print(WordsF)
	directory_search(str(windows_pwd), DirF, LogF, WordsF)

# try:
# 	r=open(str(windows_pwd) +"\\var\\log\\fmos")
# 	#r=open('sm-diagpkg.json')

# 	device_totals=[]
# 	device_type3 = list(set(device_type2))
# 	# for line in device_type2:
# 	# 	print(line)


# 	for item in device_type3:		
# 		x=device_type2.count(item)
# 		emptylist=[] 		
# 		emptylist.append(item)
# 		emptylist.append(x)
# 		device_totals.append(emptylist)


# 	device_totals.sort(key = lambda x: x[1])
# 	print(tabulate(device_totals, headers=['Device', 'Count']))

# 	print()
# 	print("Closing Check for Off Box Backup")
# 	print(formatter)

# except Exception as e:
# 	print(e)
# 	print("Error Checking for Off Box Backup")
# 	print()
# 	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# temp-files.txt

print("Checking temp-files")
print()
try:

	f = open(str(windows_pwd) +"\\temp-files.txt")
	temp_files = f.readlines()



	for line in temp_files:
		if 'devpack_pylib' in line:
			temp_files_count += 1

	print('Number temp files: '+str(temp_files_count))

	file_size = os.path.getsize(str(windows_pwd) +"\\temp-files.txt")
	file_size_mb=file_size/1000000
	print("Folder Size is:", file_size_mb, "Mb")

	print()
	print("Closing Check for temp-files")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking for temp-files")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# retrievalReport Review

print("Opening retrievalReport")
print()
try:
	if os.path.exists(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\retrievalReport.txt"):

		f = open(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\retrievalReport.txt", encoding="utf8")
		retrievalReport = f.readlines()
		Sumarry_retrievalReport=('PASS -- No Failed Retrievals found in the retrievalReport')
		for line in retrievalReport:

			if 'Failed' in line:
				Sumarry_retrievalReport=('WARNING -- Failed Retrievals found in the retrievalReport')
				
		
		print(Sumarry_retrievalReport)

	else:
		Sumarry_retrievalReport="N/A -- retrievalReport Check not performed. No DC element to MKDIAG"
		print(Sumarry_retrievalReport)

	

	print()
	print("Closing Check for retrievalReport")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking for retrievalReport")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# # syslogMessagesReport Review

# print("Opening syslogMessagesReport")
# print()
# try:
# 	if os.path.exists(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\syslogMessagesReport.txt"):

# 		f = open(str(windows_pwd) +"\\var\\log\\firemon\\dc\\reports\\syslogMessagesReport.txt", encoding="utf8")
# 		syslogMessagesReport = f.readlines()
# 		Sumarry_retrievalReport=('PASS -- No Failed Retrievals found in the syslogMessagesReport')
# 		for line in syslogMessagesReport:
# 			'''
# 			add the line to a new file list until  Unrecognized Devices: is seen then break
# 			go though that new list and drop pop every line  without device id slash
# 			get the device ID and other total from any line where the last number is not 0
# 			'''
# 			if 'Failed' in line:
# 				Sumarry_retrievalReport=('WARNING -- Failed Retrievals found in the syslogMessagesReport')
				
		
# 		print(Sumarry_retrievalReport)

# 	else:
# 		Sumarry_retrievalReport="N/A -- retrievalReport Check not performed. No DC element to MKDIAG"
# 		print(Sumarry_retrievalReport)	

	

# 	print()
# 	print("Closing Check for syslogMessagesReport")
# 	print(formatter)

# except Exception as e:
# 	print(e)
# 	print("Error Checking for syslogMessagesReport")
# 	print()
# 	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# FireMon Data

print("Opening FireMon-Data.txt File")
print()
try:
	if os.path.exists(str(windows_pwd) +"\\firemon-data.txt"):
		todaysdate2ago=datetime.today() - timedelta(days=12)
		todaysdate2ago=str(todaysdate2ago).split(' ')[0]
		print('Looking for a Backup from '+todaysdate2ago)
		print()
		backuplist=[]
		backuplistsm=[]
		f = open("firemon-data.txt")
		retrievalReport2 = f.readlines()

		
		Current_Backup_Summary=('WARNING -- No Backup Found With Current Date Stamp')
		for line in retrievalReport2:

			if '.backup' in line:
				backuplist.append(line)
				if todaysdate2ago in line:
					Current_Backup_Summary=('PASS -- A Backup Found With Current Date Stamp')
					print(line)
				

		print(Current_Backup_Summary)



				
		# for nex in backuplist[-10:]:
		# 	nex=nex.split('fmbackup')[1].split('+')[0].split('	')

		# 	nex.pop(0)
		# 	print(nex)
		# 	if todaysdate2ago in nex:
		# 		print(nex)





	else:
		print()

	

	print()
	print("Closing FireMon-Data.txt File")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Checking FireMon-Data.txt File")
	print()
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# Syslog information

# print("Checking for Off Box Backup")
# print()
# try:

# 	r=open('sm-diagpkg.json')

	

# 	print()
# 	print("Closing Check for Off Box Backup")
# 	print(formatter)

# except Exception as e:
# 	print(e)
# 	print("Error Checking for Off Box Backup")
# 	print()
# 	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

# Spacer

print()
print()
print()
print()
print()
print()
print()

#======================================================================================================================================================
#======================================================================================================================================================

# Summary Section


print(formatter)
print()


print("Summary Section")
print(formatter)

 
print()
print('General Info:')
print()

summary_list=[]
summary_list2=[]


print("Server Hostname: " + str(hostname))
summary_list.append("Server Hostname: " + str(hostname))


print("FMOS Version: " + str(FMOSv))
summary_list.append("FMOS Version: " + str(FMOSv))


print("Server Role: "+str(device))
summary_list.append("Server Role: "+str(device))


print("Uptime: " + str(Device_Uptime_host_days) + ' day(s)')
summary_list.append("Uptime: " + str(Device_Uptime_host_days) + ' day(s)')


if Check_Installed_DPs =="N/A -- Device Packs Check not performed. No DC element to MKDIAG":
	pass
else:
	print("Devices found (deviceInventoryReport): " + str(int(Total_Devices)))
	

if device=="DC only":
	pass
else:
	print("Devices found (sm_diagpkg): " + str(device_count))
	

print("MKDIAG Created: " + str(mkdiag_created))
summary_list.append("MKDIAG Created: " + str(mkdiag_created))

print()
print(formatter)


#======================================================================================================================================================
#======================================================================================================================================================

# Writing Checks Carried Out
 
print()
print('Checks carried out:')
print()

try:
	print(Check_Installed_DPs)
	summary_list2.append(Check_Installed_DPs)
except:
	pass


try:
	print(Drops_Summary)
	summary_list2.append(Drops_Summary)
except:
	pass


try:
	print(bad_syslog_summary)
	summary_list2.append(bad_syslog_summary)
except:
	pass


try:
	print(Core_Summary)
	summary_list2.append(Core_Summary)
except:
	pass


try:
	print(top_memory_summary)
	summary_list2.append(top_memory_summary)
except:
	pass


try:
	print(top_swap_summary)
	summary_list2.append(top_swap_summary)
except:
	pass


try:
	print(summary_cpu_text)
	summary_list2.append(summary_cpu_text)
except:
	pass


try:
	print(Backup_Summary)
	summary_list2.append(Backup_Summary)
except:
	pass


if device=="DC only":
	pass
else:
	print(License_Summary)
	summary_list2.append(License_Summary)
	


if device=="DC only":
	pass
else:
	print(Norm_Ret_Summary)
	summary_list2.append(Norm_Ret_Summary)


try:
	print(Sumarry_Cert)
	summary_list2.append(Sumarry_Cert)
except:
	pass


try:
	print(Summary_Dir_Check)
	summary_list2.append(Summary_Dir_Check)
except:
	pass


try:
	print(Check_Device_Uptime)
	summary_list2.append(Check_Device_Uptime)
except:
	pass


try:
	print(Sumarry_CPU)
	summary_list2.append(Sumarry_CPU)
except:
	pass


try:
	print(Sumarry_retrievalReport)
	summary_list2.append(Sumarry_retrievalReport)
except:
	pass


print(formatter)


#======================================================================================================================================================
#======================================================================================================================================================

# Writing to Excel

print("Writing to Excel: ")
print()
print('File Name: ' + str(hostname.split('.')[0])+'.xlsx')
try:

	#for item in summary_list:
	for i, item in enumerate(summary_list):
		ws.cell(row=i+2, column=1).value = item
	print('General Info ....written successfully')

	for i, item in enumerate(summary_list2):
		ws.cell(row=i+3, column=3).value = item
	print('Checks Carried Out ....written successfully')

	for i, item in enumerate(cpu_info_xls):
		ws.cell(row=i+4, column=6).value = item
	print('CPU Summary ....written successfully')



	print()
	print("Closing Writing to Excel")
	print()
	print("\U0001F40D")
	print(formatter)

except Exception as e:
	print(e)
	print("Error Writing to Excel")
	print()
	print()
	print("\U0001F40D")
	print(formatter)

#======================================================================================================================================================
#======================================================================================================================================================

#Save and Go!

wb.save(pwd + '\\script_output\\'+str(hostname.split('.')[0])+'.xlsx')
Summary_LoggingFile.close()


print()

print()
directory_search_output_analysis=[]
for line in directory_search_output:
	
	line=line.split('fmos.health')[1]
	line=line[1::]
	if 'checks.dc 'in line:
		line=line.split('checks.dc ')[1]
	else:
		pass
	line=line.split(' ')
	line.remove(line[0])
	line.remove(line[0])
	s = " "
	line = s.join(line)
	if line not in directory_search_output_analysis:
		directory_search_output_analysis.append(line)
	else:
		pass

	

for nicks in directory_search_output_analysis:
	print(nicks)


exit()

